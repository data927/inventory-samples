from __future__ import annotations

import math
import pandas as pd
import re

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE, WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BUCKETS: list[tuple[int, str]] = [
    (1, "Product & Engineering"),
    (2, "Customer & Sales"),
    (3, "Strategy & Planning"),
    (4, "Financial & Legal"),
    (5, "Operations & HR"),
    (6, "Marketing"),
    (7, "Meeting Notes & Internal Comms"),
]


def bucket_number_for_sheet_title(sheet_title: str) -> int | None:
    """Return 1–7 if ``sheet_title`` matches a bucket tab name used in workbooks, else None.

    Matches the titles produced by :func:`write_company_inventory_workbook` (plain
    bucket name, truncated to 31 chars) and the disambiguated ``\"{n}. {name}\"`` form.
    """
    title = (sheet_title or "").strip()
    if not title:
        return None
    for num, name in BUCKETS:
        primary = str(name)[:31]
        if title == primary:
            return num
        fallback = f"{num}. {name}"[:31]
        if title == fallback:
            return num
    return None

DESCRIPTION_COL_CANDIDATES = [
    "description",
    "what it includes",
    "item",
    "details",
    "notes",
    "content",
    "data",
]
SOURCE_COL_CANDIDATES = [
    "source",
    "where to find it",
    "system",
    "tool",
    "platform",
    "location",
]


def _normalize(value: object) -> str:
    return str(value).strip().lower()


def detect_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    """Return the first column matching any candidate (exact, then substring)."""
    cols_normalized = {_normalize(c): c for c in df.columns}
    for cand in candidates:
        if cand in cols_normalized:
            return cols_normalized[cand]
    for cand in candidates:
        for norm, original in cols_normalized.items():
            if cand in norm:
                return original
    return None


def load_inventory(
    path: str,
    sheet_name: str | None = None,
    description_column: str | None = None,
    source_column: str | None = None,
) -> tuple[pd.DataFrame, str, str | None]:
    """Load the inventory; return (df_with_row_id, description_col, source_col)."""
    if path.lower().endswith(".csv"):
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path, sheet_name=sheet_name or 0)

    if description_column is None:
        description_column = detect_column(df, DESCRIPTION_COL_CANDIDATES)
    if description_column is None:
        # last resort: first text column
        for col in df.columns:
            if df[col].dtype == object:
                description_column = col
                break
    if description_column is None:
        raise ValueError(
            "Could not identify a description column. "
            "Pass description_column explicitly when calling the tool."
        )

    if source_column is None:
        source_column = detect_column(df, SOURCE_COL_CANDIDATES)

    if "row_id" not in df.columns:
        df.insert(0, "row_id", range(1, len(df) + 1))

    return df, description_column, source_column


_HYPERLINK_COLS: frozenset[str] = frozenset({"Drive URL"})


def _write_sheet(ws, frame: pd.DataFrame) -> None:
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    body_font = Font(name="Arial")
    link_font = Font(name="Arial", color="0563C1", underline="single")
    body_align = Alignment(vertical="top", wrap_text=True)

    cols = list(frame.columns)
    hyperlink_col_indices: set[int] = {
        i for i, c in enumerate(cols, start=1) if c in _HYPERLINK_COLS
    }

    # Column widths from sample — computed before streaming rows
    for i, col in enumerate(cols, start=1):
        sample = frame[col].astype(str).tolist()[:200] if len(frame) else []
        max_len = max([len(str(col))] + [len(str(v)) for v in sample]) if sample else len(str(col))
        ws.column_dimensions[get_column_letter(i)].width = min(max(14, max_len + 2), 60)
    ws.freeze_panes = "A2"

    def _clean_cell_value(v):
        if v is None:
            return None
        try:
            if pd.isna(v):
                return None
        except TypeError:
            pass
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        if isinstance(v, str):
            return ILLEGAL_CHARACTERS_RE.sub("", v)
        return v

    # Header row
    header_cells = []
    for col_name in cols:
        c = WriteOnlyCell(ws, value=col_name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        header_cells.append(c)
    ws.append(header_cells)

    # Body rows — single pass, formatting applied inline
    for record in frame.itertuples(index=False):
        row_cells = []
        for col_idx, v in enumerate(list(record), start=1):
            val = _clean_cell_value(v)
            c = WriteOnlyCell(ws, value=val)
            if col_idx in hyperlink_col_indices:
                url = str(val or "").strip()
                if url.startswith("http"):
                    c.hyperlink = url
                    c.font = link_font
                else:
                    c.font = body_font
                    c.alignment = body_align
            else:
                c.font = body_font
                c.alignment = body_align
            row_cells.append(c)
        ws.append(row_cells)


def write_segmented_workbook(df: pd.DataFrame, output_path: str) -> None:
    """Write Master + 7 bucket sheets + Summary to output_path."""
    wb = Workbook(write_only=True)

    ws_master = wb.create_sheet("Master")
    _write_sheet(ws_master, df)

    for num, name in BUCKETS:
        bucket_df = df[df["bucket_number"] == num].copy()
        sheet_title = f"{num}. {name}"[:31]  # Excel's 31-char sheet name limit
        ws = wb.create_sheet(sheet_title)
        _write_sheet(ws, bucket_df)

    # Summary
    ws_summary = wb.create_sheet("Summary")
    summary_cols = ["Bucket #", "Bucket Name", "Items", "High Confidence", "Low Confidence"]
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    header_align = Alignment(horizontal="left", vertical="center")
    body_font = Font(name="Arial")

    for i in range(1, len(summary_cols) + 1):
        ws_summary.column_dimensions[get_column_letter(i)].width = 24
    ws_summary.freeze_panes = "A2"

    header_cells = []
    for col_name in summary_cols:
        c = WriteOnlyCell(ws_summary, value=col_name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        header_cells.append(c)
    ws_summary.append(header_cells)

    for num, name in BUCKETS:
        bucket_df = df[df["bucket_number"] == num]
        high = int((bucket_df.get("confidence", pd.Series([], dtype=str)) == "high").sum())
        low = int((bucket_df.get("confidence", pd.Series([], dtype=str)) == "low").sum())
        row_cells = []
        for val in [num, name, len(bucket_df), high, low]:
            c = WriteOnlyCell(ws_summary, value=val)
            c.font = body_font
            row_cells.append(c)
        ws_summary.append(row_cells)

    total_row = len(BUCKETS) + 2
    total_cells = []
    for i, val in enumerate([
        "Total", "", f"=SUM(C2:C{total_row - 1})",
        f"=SUM(D2:D{total_row - 1})", f"=SUM(E2:E{total_row - 1})"
    ], start=1):
        c = WriteOnlyCell(ws_summary, value=val)
        c.font = Font(name="Arial", bold=True)
        total_cells.append(c)
    ws_summary.append(total_cells)

    wb.save(output_path)


def evidence_display_dataframe(evidence_df: pd.DataFrame) -> pd.DataFrame:
    """Build the human-facing Evidence table from internal ``evidence_df`` columns."""
    edf = evidence_df.copy()

    def _str_col(name: str, default: str = "") -> pd.Series:
        if name not in edf.columns:
            return pd.Series(default, index=edf.index, dtype="string")
        s = edf[name].astype("string")
        return s.fillna(default).replace({pd.NA: default, "<NA>": default})

    def _size_human(n: object) -> str:
        if n is None:
            return ""
        try:
            if pd.isna(n):
                return ""
        except TypeError:
            pass
        try:
            x = float(n)
        except (TypeError, ValueError):
            return ""
        if pd.isna(x) or x < 0 or not math.isfinite(x):
            return ""
        units = ["B", "KB", "MB", "GB", "TB"]
        i = 0
        while x >= 1024 and i < len(units) - 1:
            x /= 1024
            i += 1
        if i == 0:
            return f"{int(x)} {units[i]}"
        return f"{x:.1f} {units[i]}"

    if "word_count" in edf.columns:
        wc_series = pd.to_numeric(edf["word_count"], errors="coerce").fillna(0).astype("int64")
    else:
        wc_series = pd.Series(0, index=edf.index, dtype="int64")

    if "token_count" in edf.columns:
        tk_series = pd.to_numeric(edf["token_count"], errors="coerce").fillna(0).astype("int64")
    else:
        tk_series = pd.Series(0, index=edf.index, dtype="int64")

    if "page_count" in edf.columns:
        pc_series = pd.to_numeric(edf["page_count"], errors="coerce").fillna(0).astype("int64")
    else:
        pc_series = pd.Series(0, index=edf.index, dtype="int64")

    if "modality" in edf.columns:
        modality_series = _str_col("modality", "")
    else:
        modality_series = pd.Series("", index=edf.index, dtype="string")

    if "content_type" in edf.columns:
        content_type_series = _str_col("content_type", "")
    else:
        content_type_series = pd.Series("", index=edf.index, dtype="string")

    if "quality_tier" in edf.columns:
        quality_tier_series = _str_col("quality_tier", "")
    else:
        quality_tier_series = pd.Series("", index=edf.index, dtype="string")

    if "content_inventory" in edf.columns:
        content_inventory_series = _str_col("content_inventory", "")
    else:
        content_inventory_series = pd.Series("", index=edf.index, dtype="string")

    if "sub_category" in edf.columns:
        sub_category_series = _str_col("sub_category", "")
    else:
        sub_category_series = pd.Series("", index=edf.index, dtype="string")

    size_series = (
        edf["size_bytes"].map(_size_human)
        if "size_bytes" in edf.columns
        else pd.Series("", index=edf.index, dtype="string")
    )

    # New attribution + traceability columns (populated by pipeline_1tb; empty
    # string when not present so existing workbooks are unaffected).
    drive_url_series = _str_col("drive_url", "")
    owner_email_series = _str_col("owner_email", "")
    last_modified_by_series = _str_col("last_modified_by", "")
    llm_pass_series = _str_col("llm_pass", "")

    return pd.DataFrame(
        {
            "Item Name": _str_col("filename", ""),
            "File Type": _str_col("extension", ""),
            "Content Summary": _str_col("content_summary", ""),
            "PII Flag": _str_col("pii_flag", ""),
            "Size": size_series,
            "Word Count": wc_series,
            "LLM Tokens": tk_series,
            "Page Count": pc_series,
            "Modality": modality_series,
            "Content Type": content_type_series,
            "Quality tier": quality_tier_series,
            "What is contained": content_inventory_series,
            "Source": _str_col("source_guess", ""),
            "Owner": owner_email_series,
            "Last Modified By": last_modified_by_series,
            "Drive URL": drive_url_series,
            "LLM Pass": llm_pass_series,
            "Path / Subsection": _str_col("path", ""),
            "bucket_number": _str_col("bucket_number", ""),
            "bucket_name": _str_col("bucket_name", ""),
            "Sub-category": sub_category_series,
            "confidence": _str_col("confidence", ""),
            "rationale": _str_col("rationale", ""),
        }
    )


def write_company_inventory_workbook(
    *,
    inventory_df: pd.DataFrame,
    evidence_df: pd.DataFrame | None,
    output_path: str,
) -> None:
    """Write a clean multi-tab workbook for dump-derived outputs.

    Sheets:
      - Inventory: 7-row desired format
      - Overview: quick counts by bucket/confidence/source/ext plus modality/content-type/quality tier tops
      - Evidence: full evidence table (same columns as Sample Files + bucket tabs)
      - Sample Files: first N evidence rows (easy browsing)
      - One sheet per bucket: "<bucket name>" with evidence rows for that bucket
    """
    wb = Workbook()
    wb.remove(wb.active)

    ws_inventory = wb.create_sheet("Inventory")
    _write_sheet(ws_inventory, inventory_df)

    if evidence_df is not None:
        edf = evidence_df.copy()
        display = evidence_display_dataframe(edf)

        # Overview
        ws_overview = wb.create_sheet("Overview")
        overview_rows: list[dict[str, object]] = []

        # Bucket summary
        if "bucket_number" in evidence_df.columns:
            bucket_counts = (
                pd.to_numeric(evidence_df["bucket_number"], errors="coerce")
                .value_counts(dropna=False)
                .sort_index()
                .to_dict()
            )
            for k, v in bucket_counts.items():
                overview_rows.append({"Metric": "bucket_count", "Key": str(k), "Value": int(v)})

        # Confidence summary
        if "confidence" in evidence_df.columns:
            conf_counts = evidence_df["confidence"].value_counts(dropna=False).to_dict()
            for k, v in conf_counts.items():
                overview_rows.append({"Metric": "confidence_count", "Key": str(k), "Value": int(v)})

        # Source guesses
        if "source_guess" in evidence_df.columns:
            src_counts = (
                evidence_df["source_guess"]
                .fillna("")
                .astype(str)
                .replace({"nan": ""})
                .value_counts()
                .head(40)
                .to_dict()
            )
            for k, v in src_counts.items():
                key = k if k else "(blank)"
                overview_rows.append({"Metric": "source_guess_top", "Key": key, "Value": int(v)})

        # File extensions
        if "extension" in evidence_df.columns:
            ext_counts = (
                evidence_df["extension"]
                .fillna("")
                .astype(str)
                .replace({"nan": ""})
                .str.lower()
                .value_counts()
                .head(40)
                .to_dict()
            )
            for k, v in ext_counts.items():
                key = k if k else "(blank)"
                overview_rows.append({"Metric": "extension_top", "Key": key, "Value": int(v)})

        if "quality_tier" in evidence_df.columns:
            qt_counts = (
                evidence_df["quality_tier"]
                .fillna("")
                .astype(str)
                .replace({"nan": ""})
                .value_counts()
                .head(20)
                .to_dict()
            )
            for k, v in qt_counts.items():
                key = k if k else "(blank)"
                overview_rows.append({"Metric": "quality_tier_top", "Key": key, "Value": int(v)})

        if "modality" in evidence_df.columns:
            mod_counts = (
                evidence_df["modality"]
                .fillna("")
                .astype(str)
                .replace({"nan": ""})
                .value_counts()
                .head(40)
                .to_dict()
            )
            for k, v in mod_counts.items():
                key = k if k else "(blank)"
                overview_rows.append({"Metric": "modality_top", "Key": key, "Value": int(v)})

        if "content_type" in evidence_df.columns:
            ct_counts = (
                evidence_df["content_type"]
                .fillna("")
                .astype(str)
                .replace({"nan": ""})
                .value_counts()
                .head(40)
                .to_dict()
            )
            for k, v in ct_counts.items():
                key = k if k else "(blank)"
                overview_rows.append({"Metric": "content_type_top", "Key": key, "Value": int(v)})

        if "sub_category" in evidence_df.columns:
            sc_counts = (
                evidence_df["sub_category"]
                .fillna("")
                .astype(str)
                .replace({"nan": ""})
                .value_counts()
                .head(40)
                .to_dict()
            )
            for k, v in sc_counts.items():
                key = k if k else "(blank)"
                overview_rows.append({"Metric": "sub_category_top", "Key": key, "Value": int(v)})

        if overview_rows:
            _write_sheet(ws_overview, pd.DataFrame(overview_rows))
        else:
            _write_sheet(ws_overview, pd.DataFrame([{"Metric": "info", "Key": "", "Value": "No evidence provided"}]))

        # Evidence + Sample Files + bucket tabs share the same column layout (all enrichment columns).
        ws_evidence = wb.create_sheet("Evidence")
        _write_sheet(ws_evidence, display)

        ws_sample = wb.create_sheet("Sample Files")
        sample_n = 250 if len(display) >= 250 else len(display)
        _write_sheet(ws_sample, display.head(sample_n))

        # One tab per bucket with that bucket's evidence (clean browsing)
        if "bucket_number" in display.columns and "bucket_name" in display.columns:
            bn_series = pd.to_numeric(display["bucket_number"], errors="coerce")
            for num, name in BUCKETS:
                bdf = display[bn_series == num].copy()
                if len(bdf) == 0:
                    continue
                # Prefer the plain bucket name as sheet title (what users expect)
                sheet_title = str(name)[:31]
                # Ensure uniqueness if names collide after truncation
                if sheet_title in wb.sheetnames:
                    sheet_title = f"{num}. {name}"[:31]
                ws = wb.create_sheet(sheet_title)
                _write_sheet(ws, bdf)

        # Low Confidence tab — all confidence=low rows for human review
        if "confidence" in display.columns:
            low_df = display[display["confidence"].astype(str).str.strip().str.lower() == "low"].copy()
            if not low_df.empty:
                ws_low = wb.create_sheet("Low Confidence")
                _write_sheet(ws_low, low_df)

        # PII Flagged tab — rows where PII was detected
        if "PII Flag" in display.columns:
            pii_df = display[display["PII Flag"].astype(str).str.strip().str.lower() == "yes"].copy()
            if not pii_df.empty:
                ws_pii = wb.create_sheet("PII Flagged")
                _write_sheet(ws_pii, pii_df)

    wb.save(output_path)
