"""Merge multiple enriched-full xlsx files into one, appending same-named sheets row-by-row."""
import openpyxl

SOURCES = [
    "out/aiop-product-features-enriched-full.xlsx",
    "out/may23-packetai-dd-enriched-full.xlsx",
    "out/packetai-board-enriched-full.xlsx",
]
OUTPUT = "out/combined-all-enriched-full.xlsx"


def copy_row(src_ws, src_row_idx, dst_ws):
    row_data = []
    for cell in src_ws[src_row_idx]:
        row_data.append(cell.value)
    dst_ws.append(row_data)


def merge():
    # Load all workbooks
    workbooks = [openpyxl.load_workbook(f) for f in SOURCES]

    # Collect all sheet names in order of first appearance
    seen = set()
    sheet_order = []
    for wb in workbooks:
        for name in wb.sheetnames:
            if name not in seen:
                sheet_order.append(name)
                seen.add(name)

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)  # remove default empty sheet

    for sheet_name in sheet_order:
        out_ws = out_wb.create_sheet(title=sheet_name)
        header_written = False

        for wb in workbooks:
            if sheet_name not in wb.sheetnames:
                continue
            src_ws = wb[sheet_name]
            for row_idx, row in enumerate(src_ws.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    if not header_written:
                        out_ws.append(list(row))
                        header_written = True
                    # always skip header from subsequent files
                else:
                    out_ws.append(list(row))

    out_wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")

    # Verify
    vwb = openpyxl.load_workbook(OUTPUT, read_only=True)
    for sh in vwb.sheetnames:
        ws = vwb[sh]
        print(f"  {sh}: {ws.max_row} rows x {ws.max_column} cols")
    vwb.close()


if __name__ == "__main__":
    merge()
